from utils import timeit
from typing import Any

import openpyxl


@timeit
def create_excel(result: dict[str, Any]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "test"
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 100

    areas_col = 1
    cities_col = 2
    warehouses_col = 3

    headers = {
        areas_col: "Areas",
        cities_col: "Cities",
        warehouses_col: "Warehouses"
    }
    for col, header in headers.items():
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    current_row = 2
    for area, cities in result.items():
        total_warehouses = sum(len(warehouses) for warehouses in cities.values())
        top_left_cell = ws.cell(row=current_row, column=areas_col, value=area)
        top_left_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        if total_warehouses > 0:
            ws.merge_cells(
                start_row=current_row,
                start_column=areas_col,
                end_row=current_row + total_warehouses - 1,
                end_column=areas_col
            )

        for city, warehouses_of_the_city in cities.items():
            len_warehouses = len(warehouses_of_the_city)
            top_left_cell = ws.cell(row=current_row, column=cities_col, value=city)
            top_left_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            if len_warehouses > 1:
                ws.merge_cells(
                    start_row=current_row,
                    start_column=cities_col,
                    end_row=current_row + len_warehouses - 1,
                    end_column=cities_col
                )

            for warehouse in warehouses_of_the_city:
                ws.cell(row=current_row, column=warehouses_col, value=warehouse)
                current_row += 1
    current_row += 1

    wb.save('../test.xlsx')

